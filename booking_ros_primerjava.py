import io
import re
import csv
import hashlib
import unicodedata
from datetime import datetime

import pandas as pd
import numpy as np
import streamlit as st

st.set_page_config(page_title="Booking.com vs ROS PMS", layout="wide")

# ---------------------------------------------------------------------------
# Pomožne funkcije za branje in čiščenje podatkov
# ---------------------------------------------------------------------------

def normalize_name_tokens(name) -> frozenset:
    """Iz imena naredi normaliziran nabor besed (brez šumnikov, ločil, velikih
    črk) - uporabno za ujemanje ne glede na vrstni red (Ime Priimek vs
    Priimek Ime) in drobne razlike v zapisu."""
    if name is None:
        return frozenset()
    s = str(name).upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z\s]", " ", s)
    tokens = {t for t in s.split() if len(t) > 1}
    return frozenset(tokens)


def name_match_score(tokens_a: frozenset, tokens_b: frozenset) -> float:
    """Delež ujemajočih se besed glede na manjši od obeh naborov (0-1)."""
    if not tokens_a or not tokens_b:
        return 0.0
    common = tokens_a & tokens_b
    return len(common) / min(len(tokens_a), len(tokens_b))


def parse_ros_csv(file_bytes: bytes) -> pd.DataFrame:
    """Prebere ROS PMS CSV izvoz (';'-ločen, cp1250, lahko vsebuje
    večvrstična citirana polja in je na koncu lahko obrezan)."""
    text = None
    for enc in ["cp1250", "iso-8859-2", "cp1252", "utf-8"]:
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = file_bytes.decode("cp1250", errors="replace")

    # poenoti konce vrstic, da csv modul pravilno obravnava citirana polja
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    reader = csv.reader(io.StringIO(text), delimiter=";", quotechar='"')
    rows = list(reader)
    if not rows:
        return pd.DataFrame()

    header = rows[0]
    ncols = len(header)
    data = []
    for r in rows[1:]:
        if not r or all((x is None or x == "") for x in r):
            continue
        if len(r) < ncols:
            r = r + [""] * (ncols - len(r))  # zadnja vrstica je lahko obrezana
        elif len(r) > ncols:
            r = r[:ncols]
        data.append(r)

    df = pd.DataFrame(data, columns=header)
    return df


def parse_sl_date(s):
    """Pretvori slovenski datumski zapis npr. '1. 07. 2026' v pd.Timestamp."""
    if s is None:
        return pd.NaT
    s = str(s).strip()
    if not s:
        return pd.NaT
    m = re.match(r"(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})", s)
    if not m:
        return pd.NaT
    d, mo, y = m.groups()
    try:
        return pd.Timestamp(year=int(y), month=int(mo), day=int(d))
    except ValueError:
        return pd.NaT


def parse_sl_number(s):
    """Pretvori slovenski format števila ('996,6') v float."""
    if s is None:
        return None
    s = str(s).strip().replace(" ", "")
    if s == "":
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def file_hash(uploaded_file) -> str:
    uploaded_file.seek(0)
    h = hashlib.md5(uploaded_file.read()).hexdigest()
    uploaded_file.seek(0)
    return h


MONEY_COLUMNS = {
    "Promet (ROS)", "Dodatek (ROS)", "Skupaj ROS", "Final amount (BC)",
    "Razlika (Skupaj ROS - Final)", "Provizija znesek (BC)",
    "Provizija pričakovana", "Razlika provizije", "Neto znesek (Final - Provizija)",
}
PERCENT_COLUMNS = {"Provizija %"}


def build_xlsx_export(df: pd.DataFrame) -> bytes:
    """Izdela formatiran .xlsx (Arial pisava, obarvani status, denarni format
    na EUR stolpcih, zamrznjena glava, avtomatska širina stolpcev)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Primerjava"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    ok_fill = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")
    warn_fill = PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid")
    missing_fill = PatternFill(start_color="F8D0D0", end_color="F8D0D0", fill_type="solid")

def sanitize_excel_value(v):
    """Pretvori pandas/numpy vrednosti v čiste Python tipe, ki jih openpyxl
    zna zapisati (numpy.bool_, numpy.int64, numpy.float64, NaN, NaT ...)."""
    if v is None:
        return None
    if isinstance(v, np.generic):
        v = v.item()  # numpy skalar -> native python (int/float/bool)
    if isinstance(v, pd.Timestamp):
        return None if pd.isna(v) else v.to_pydatetime()
    if isinstance(v, float) and pd.isna(v):
        return None
    if isinstance(v, bool):
        return v
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (frozenset, set)):
        return ", ".join(sorted(str(x) for x in v))
    if not isinstance(v, (str, int, float, bool, datetime)):
        return str(v)
    return v


def build_xlsx_export(df: pd.DataFrame) -> bytes:
    """Izdela formatiran .xlsx (Arial pisava, obarvani status, denarni format
    na EUR stolpcih, zamrznjena glava, avtomatska širina stolpcev)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Primerjava"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    ok_fill = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")
    warn_fill = PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid")
    missing_fill = PatternFill(start_color="F8D0D0", end_color="F8D0D0", fill_type="solid")

    columns = list(df.columns)
    ws.append(columns)
    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    status_col_idx = columns.index("Status") + 1 if "Status" in columns else None

    for _, row in df.iterrows():
        values = [sanitize_excel_value(row[col]) for col in columns]
        ws.append(values)
        r = ws.max_row
        status_val = str(row.get("Status", ""))
        if status_val.startswith("❌"):
            row_fill = missing_fill
        elif "Ne štima" in status_val:
            row_fill = warn_fill
        elif status_val.startswith("✅") or "OK" in status_val:
            row_fill = ok_fill
        else:
            row_fill = None

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = normal_font
            if row_fill is not None:
                cell.fill = row_fill
            if col in MONEY_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "€"'
            elif col in PERCENT_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # avtomatska (približna) širina stolpcev
    for col_idx, col in enumerate(columns, start=1):
        max_len = len(str(col))
        for val in df[col]:
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Obdelava podatkov
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def load_booking_df(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_csv(io.BytesIO(file_bytes), encoding="utf-8")
    df.columns = [c.strip() for c in df.columns]
    df["Reservation number"] = df["Reservation number"].astype(str).str.strip()
    for col in ["Arrival", "Departure"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    for col in ["Original amount", "Final amount", "Commission amount", "Commission %"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["Nights_BC"] = (df["Departure"] - df["Arrival"]).dt.days
    return df


@st.cache_data(show_spinner=False)
def load_ros_df(file_bytes: bytes) -> pd.DataFrame:
    raw = parse_ros_csv(file_bytes)
    if raw.empty:
        return raw

    raw["Zacetek_dt"] = raw["Začetek"].apply(parse_sl_date)
    raw["Konec_dt"] = raw["Konec"].apply(parse_sl_date)
    raw["Cena_num"] = raw["Cena"].apply(parse_sl_number)
    raw["Promet_num"] = raw["Promet"].apply(parse_sl_number)
    raw["Referenca"] = raw["Referenca"].astype(str).str.strip()

    rows = []
    for rez, g in raw.groupby("Rezervacija", dropna=False):
        # Promet je včasih podvojen na več vrsticah iste rezervacije
        # (isti znesek, različne vrstice/dolocila) - v tem primeru ga ne
        # seštevamo, sicer pa seštejemo, ker gre za dejansko ločene postavke.
        unique_promet = pd.Series(g["Promet_num"].round(4)).dropna().unique()
        if len(unique_promet) == 0:
            total_promet = 0.0
        elif len(unique_promet) == 1:
            total_promet = float(unique_promet[0])
        else:
            total_promet = float(unique_promet.sum())

        cena_values = pd.Series(g["Cena_num"]).dropna().unique()
        cena_display = cena_values[0] if len(cena_values) == 1 else (
            ", ".join(f"{v:g}" for v in cena_values) if len(cena_values) else None
        )

        zacetek = g["Zacetek_dt"].min()
        konec = g["Konec_dt"].max()
        referenca = next((r for r in g["Referenca"] if r), "")
        naziv_skupine = next((v for v in g.get("Naziv skupine", []) if v), "")
        gost = next((v for v in g.get("IME_GOSTA", []) if v), "")
        prostor = next((v for v in g.get("Naziv tipa", []) if v), "")

        rows.append({
            "Rezervacija_ROS": rez,
            "Referenca": referenca,
            "Naziv skupine": naziv_skupine,
            "Gost_ROS": gost,
            "Tip prostora": prostor,
            "Zacetek_ROS": zacetek,
            "Konec_ROS": konec,
            "Nights_ROS": (konec - zacetek).days if pd.notna(zacetek) and pd.notna(konec) else None,
            "Cena_ROS": cena_display,
            "Promet_ROS": round(total_promet, 2),
            "St_vrstic": len(g),
        })

    return pd.DataFrame(rows)


def match_reservations(bc_df: pd.DataFrame, ros_df: pd.DataFrame, name_match_threshold: float = 0.6):
    """Poveže vsako Booking.com rezervacijo z ROS rezervacijo v dveh korakih:
    1. po Referenci (natančno ujemanje Reservation number = Referenca)
    2. za tiste, ki jih ni bilo mogoče najti po referenci, poskusi ujemanje
       po imenu/priimku gosta (normalizirano, brez šumnikov, ne glede na
       vrstni red besed) med še nezasedenimi ROS vrsticami.
    Vrne DataFrame enake dolžine kot bc_df s stolpci '_ros_match_idx' in
    'Nacin_ujemanja' ('referenca' / 'ime' / None)."""

    bc_df = bc_df.reset_index(drop=True).copy()
    ros_df = ros_df.reset_index(drop=True).copy()

    bc_df["_name_tokens"] = bc_df["Guest name"].apply(normalize_name_tokens)
    ros_df["_name_tokens"] = ros_df["Gost_ROS"].apply(normalize_name_tokens)

    ref_to_indices = {}
    for idx, ref in ros_df["Referenca"].items():
        if ref:
            ref_to_indices.setdefault(ref, []).append(idx)

    match_idx = [None] * len(bc_df)
    match_type = [None] * len(bc_df)
    used_ros = set()

    # 1. korak: ujemanje po referenci
    for bidx, resnum in bc_df["Reservation number"].items():
        candidates = [c for c in ref_to_indices.get(resnum, []) if c not in used_ros]
        if candidates:
            match_idx[bidx] = candidates[0]
            match_type[bidx] = "referenca"
            used_ros.add(candidates[0])

    # 2. korak: ujemanje po imenu za tiste, ki jih referenca ni razrešila
    for bidx in bc_df.index:
        if match_idx[bidx] is not None:
            continue
        b_tokens = bc_df.at[bidx, "_name_tokens"]
        if not b_tokens:
            continue
        best_idx, best_score = None, 0.0
        for ridx in ros_df.index:
            if ridx in used_ros:
                continue
            r_tokens = ros_df.at[ridx, "_name_tokens"]
            score = name_match_score(b_tokens, r_tokens)
            if score > best_score:
                best_score, best_idx = score, ridx
        if best_idx is not None and best_score >= name_match_threshold:
            match_idx[bidx] = best_idx
            match_type[bidx] = "ime"
            used_ros.add(best_idx)

    bc_df["_ros_match_idx"] = match_idx
    bc_df["Nacin_ujemanja"] = match_type
    return bc_df, ros_df


def build_comparison(bc_df: pd.DataFrame, ros_df: pd.DataFrame, dodatki: dict, tolerance: float) -> pd.DataFrame:
    bc_matched, ros_indexed = match_reservations(bc_df, ros_df)

    ros_cols = [
        "Rezervacija_ROS", "Referenca", "Naziv skupine", "Gost_ROS", "Tip prostora",
        "Zacetek_ROS", "Konec_ROS", "Nights_ROS", "Cena_ROS", "Promet_ROS", "St_vrstic",
    ]

    def get_field(ridx, col):
        if ridx is None or pd.isna(ridx):
            return None
        return ros_indexed.at[int(ridx), col]

    for col in ros_cols:
        bc_matched[col] = bc_matched["_ros_match_idx"].apply(lambda ridx, c=col: get_field(ridx, c))

    merged = bc_matched

    merged["Dodatek_ROS"] = merged["Reservation number"].map(
        lambda r: dodatki.get(r, 0.0)
    ).fillna(0.0)

    merged["V_ROS_najdeno"] = merged["_ros_match_idx"].notna()

    merged["Skupaj_ROS"] = merged["Promet_ROS"].fillna(0.0) + merged["Dodatek_ROS"]

    merged["Razlika_znesek"] = merged["Skupaj_ROS"] - merged["Final amount"]
    # Znesek je problematičen SAMO, če je ROS znesek NIŽJI od Final amount
    # (torej v ROS manjka promet). Če je ROS znesek VIŠJI, to praviloma pomeni,
    # da je gost podaljšal bivanje in razliko ŽE PLAČAL neposredno v ROS
    # (mimo Booking.com) - gost ničesar ne dolguje, plačilo je že prejeto.
    # Booking.com od te razlike ne zaračuna provizije, ker ni šla skozi njih.
    merged["Znesek_OK"] = merged["V_ROS_najdeno"] & (merged["Razlika_znesek"] >= -tolerance)
    merged["Podaljsano_placano_v_ROS"] = merged["V_ROS_najdeno"] & (merged["Razlika_znesek"] > tolerance)

    merged["Datumi_OK"] = (
        merged["V_ROS_najdeno"]
        & (merged["Arrival"] == merged["Zacetek_ROS"])
        & (merged["Departure"] == merged["Konec_ROS"])
    )

    merged["Provizija_pricakovana"] = (merged["Final amount"] * merged["Commission %"] / 100.0).round(2)
    merged["Provizija_razlika"] = merged["Provizija_pricakovana"] - merged["Commission amount"]
    merged["Provizija_OK"] = merged["Provizija_razlika"].abs() <= tolerance

    merged["Neto_znesek_Booking"] = merged["Final amount"] - merged["Commission amount"]

    def status(row):
        if not row["V_ROS_najdeno"]:
            return "❌ Ni v ROS (niti po referenci niti po imenu)"
        probs = []
        if not row["Datumi_OK"]:
            probs.append("datumi")
        if not row["Znesek_OK"]:
            probs.append("znesek (manjka promet v ROS)")
        if not row["Provizija_OK"]:
            probs.append("provizija")
        prefix = "🟡 (ujemanje po imenu, ne po referenci) " if row["Nacin_ujemanja"] == "ime" else ""
        if probs:
            return prefix + "⚠️ Ne štima: " + ", ".join(probs)
        if row["Podaljsano_placano_v_ROS"]:
            return prefix + "✅ OK (podaljšano bivanje - gost je razliko že plačal v ROS, brez provizije)"
        return prefix + "✅ OK"

    merged["Status"] = merged.apply(status, axis=1)
    return merged


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.title("📊 Primerjava: Booking.com statement vs ROS PMS")
st.caption(
    "Naloži Booking.com 'reservation statements overview' (nosilna datoteka) in "
    "ROS PMS izvoz rezervacij. Aplikacija poveže rezervacije preko polja "
    "**Referenca** (ROS) = **Reservation number** (Booking.com), primerja datume, "
    "izračuna promet in preveri provizijo."
)

with st.sidebar:
    st.header("1. Nalaganje datotek")
    bc_file = st.file_uploader("Booking.com statement (reservation_statements_overview...)", type=["csv"])
    ros_file = st.file_uploader("ROS PMS izvoz rezervacij (SEZNAM_REZERVACIJ...)", type=["csv"])

    st.header("2. Nastavitve")
    tolerance = st.number_input(
        "Toleranca za ujemanje zneskov (EUR)",
        min_value=0.0, max_value=50.0, value=0.02, step=0.01,
        help="Znotraj te razlike se znesek/provizija šteje kot ujemajoč.",
    )
    show_only_mismatch = st.checkbox("Prikaži samo neujemajoče rezervacije", value=False)

if not bc_file or not ros_file:
    st.info("⬅️ V stranski vrstici naloži obe CSV datoteki, da se prikaže primerjava.")
    st.stop()

bc_bytes = bc_file.getvalue()
ros_bytes = ros_file.getvalue()

try:
    bc_df = load_booking_df(bc_bytes)
except Exception as e:
    st.error(f"Napaka pri branju Booking.com datoteke: {e}")
    st.stop()

try:
    ros_df = load_ros_df(ros_bytes)
except Exception as e:
    st.error(f"Napaka pri branju ROS PMS datoteke: {e}")
    st.stop()

if ros_df.empty:
    st.error("ROS PMS datoteka je prazna ali je ni bilo mogoče prebrati.")
    st.stop()

# ključ za session_state, da se dodatki ne pobrišejo ob vsakem ponovnem zagonu,
# a se resetirajo, če se naložita nove datoteke
data_key = f"dodatki_{file_hash(bc_file)}_{file_hash(ros_file)}"
if data_key not in st.session_state:
    st.session_state[data_key] = {}

dodatki = st.session_state[data_key]

comparison = build_comparison(bc_df, ros_df, dodatki, tolerance)

# ---------------------------------------------------------------------------
# Povzetek
# ---------------------------------------------------------------------------

total_bc = len(comparison)
found_in_ros = int(comparison["V_ROS_najdeno"].sum())
not_found = total_bc - found_in_ros
ok_count = int(comparison["Status"].str.startswith("✅").sum())
extended_count = int(comparison["Podaljsano_placano_v_ROS"].sum())

c1, c2, c3, c4 = st.columns(4)
c1.metric("Rezervacij v Booking.com", total_bc)
c2.metric("Najdenih v ROS", found_in_ros)
c3.metric("✅ Ujemajo se / OK", ok_count)
c4.metric("⚠️ Ne štima / manjka", total_bc - ok_count)

if extended_count:
    st.caption(
        f"ℹ️ Pri {extended_count} rezervacijah je ROS znesek višji od Final amount "
        "(gost je verjetno podaljšal bivanje in razliko že plačal neposredno v ROS, "
        "mimo Booking.com) - to je označeno kot OK, gost ničesar ne dolguje, "
        "provizija se od te razlike ne obračunava."
    )

st.divider()

# ---------------------------------------------------------------------------
# Urejevalna tabela - vnos dodatka na rezervacijo
# ---------------------------------------------------------------------------

st.subheader("💶 Dodatek na ROS znesek (poljubno polje)")
st.caption(
    "Če ima rezervacija v ROS dodatne postavke, ki ne spadajo v osnovni Promet "
    "(npr. turistična taksa, doplačila ipd.), jih vnesi tukaj po rezervaciji - "
    "znesek se bo prištel k Prometu iz ROS in ponovno preveril z Final amount."
)

editable_cols = ["Reservation number", "Guest name", "Promet_ROS"]
editable_df = comparison[editable_cols].copy()
editable_df["Dodatek_ROS"] = editable_df["Reservation number"].map(lambda r: dodatki.get(r, 0.0))
editable_df = editable_df.rename(columns={
    "Reservation number": "Booking rezervacija",
    "Guest name": "Gost",
    "Promet_ROS": "Promet (ROS, iz datoteke)",
})

edited = st.data_editor(
    editable_df,
    key=f"editor_{data_key}",
    hide_index=True,
    use_container_width=True,
    disabled=["Booking rezervacija", "Gost", "Promet (ROS, iz datoteke)"],
    column_config={
        "Dodatek_ROS": st.column_config.NumberColumn(
            "Dodatek ROS (EUR)", help="Poljuben dodaten znesek za to rezervacijo", step=0.01, format="%.2f"
        ),
    },
)

# shrani spremembe nazaj v session_state in ponovno izračunaj primerjavo
changed = False
for _, r in edited.iterrows():
    key = r["Booking rezervacija"]
    val = float(r["Dodatek_ROS"]) if pd.notna(r["Dodatek_ROS"]) else 0.0
    if dodatki.get(key, 0.0) != val:
        dodatki[key] = val
        changed = True

if changed:
    st.session_state[data_key] = dodatki
    comparison = build_comparison(bc_df, ros_df, dodatki, tolerance)
    ok_count = int(comparison["Status"].str.startswith("✅").sum())

st.divider()

# ---------------------------------------------------------------------------
# Glavna primerjalna tabela
# ---------------------------------------------------------------------------

st.subheader("📋 Primerjalna tabela")

display_df = comparison.copy()
display_df["Arrival"] = display_df["Arrival"].dt.strftime("%d.%m.%Y")
display_df["Departure"] = display_df["Departure"].dt.strftime("%d.%m.%Y")
display_df["Zacetek_ROS"] = pd.to_datetime(display_df["Zacetek_ROS"], errors="coerce").dt.strftime("%d.%m.%Y")
display_df["Konec_ROS"] = pd.to_datetime(display_df["Konec_ROS"], errors="coerce").dt.strftime("%d.%m.%Y")
display_df["Nacin_ujemanja"] = display_df["Nacin_ujemanja"].map({
    "referenca": "po referenci",
    "ime": "po imenu (preveri ročno!)",
}).fillna("ni najdeno")

cols_map = {
    "Reservation number": "Rezervacija (Booking.com)",
    "Guest name": "Gost",
    "Status": "Status",
    "Nacin_ujemanja": "Način ujemanja",
    "Rezervacija_ROS": "Rezervacija (ROS)",
    "Gost_ROS": "Gost (ROS)",
    "Arrival": "Prihod (BC)",
    "Departure": "Odhod (BC)",
    "Nights_BC": "Noči (BC)",
    "Zacetek_ROS": "Začetek (ROS)",
    "Konec_ROS": "Konec (ROS)",
    "Nights_ROS": "Noči (ROS)",
    "Datumi_OK": "Datumi OK?",
    "Cena_ROS": "Cena/dan (ROS)",
    "Promet_ROS": "Promet (ROS)",
    "Dodatek_ROS": "Dodatek (ROS)",
    "Skupaj_ROS": "Skupaj ROS",
    "Final amount": "Final amount (BC)",
    "Razlika_znesek": "Razlika (Skupaj ROS - Final)",
    "Znesek_OK": "Znesek OK?",
    "Podaljsano_placano_v_ROS": "Podaljšano bivanje - že plačano v ROS?",
    "Commission %": "Provizija %",
    "Commission amount": "Provizija znesek (BC)",
    "Provizija_pricakovana": "Provizija pričakovana",
    "Provizija_razlika": "Razlika provizije",
    "Provizija_OK": "Provizija OK?",
    "Neto_znesek_Booking": "Neto znesek (Final - Provizija)",
}
display_df = display_df[list(cols_map.keys())].rename(columns=cols_map)

if show_only_mismatch:
    display_df = display_df[display_df["Status"] != "✅ OK"]

st.dataframe(
    display_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Promet (ROS)": st.column_config.NumberColumn(format="%.2f €"),
        "Dodatek (ROS)": st.column_config.NumberColumn(format="%.2f €"),
        "Skupaj ROS": st.column_config.NumberColumn(format="%.2f €"),
        "Final amount (BC)": st.column_config.NumberColumn(format="%.2f €"),
        "Razlika (Skupaj ROS - Final)": st.column_config.NumberColumn(format="%.2f €"),
        "Provizija znesek (BC)": st.column_config.NumberColumn(format="%.2f €"),
        "Provizija pričakovana": st.column_config.NumberColumn(format="%.2f €"),
        "Razlika provizije": st.column_config.NumberColumn(format="%.2f €"),
        "Neto znesek (Final - Provizija)": st.column_config.NumberColumn(format="%.2f €"),
    },
)

xlsx_export = build_xlsx_export(display_df)
st.download_button(
    "⬇️ Prenesi primerjalno tabelo (Excel .xlsx)",
    data=xlsx_export,
    file_name="primerjava_booking_ros.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.divider()

# ---------------------------------------------------------------------------
# Neujemajoče / manjkajoče rezervacije iz ROS strani (kanal BOOKING.COM,
# a brez ujemajočega Reservation number v Booking.com datoteki)
# ---------------------------------------------------------------------------

st.subheader("🔍 ROS rezervacije s kanalom BOOKING.COM brez ujemanja v Booking.com statementu")

bc_numbers = set(bc_df["Reservation number"])
ros_booking_channel = ros_df[
    ros_df["Naziv skupine"].astype(str).str.upper().str.contains("BOOKING", na=False)
    | ros_df["Referenca"].astype(str).str.match(r"^\d{9,10}$", na=False)
]
missing_in_bc = ros_booking_channel[~ros_booking_channel["Referenca"].isin(bc_numbers)]

if missing_in_bc.empty:
    st.success("Ni ROS rezervacij s kanalom Booking.com, ki bi manjkale v Booking.com statementu.")
else:
    show = missing_in_bc[[
        "Rezervacija_ROS", "Referenca", "Gost_ROS", "Zacetek_ROS", "Konec_ROS", "Promet_ROS"
    ]].rename(columns={
        "Rezervacija_ROS": "Rezervacija (ROS)",
        "Referenca": "Referenca (Booking št.)",
        "Gost_ROS": "Gost",
        "Zacetek_ROS": "Začetek",
        "Konec_ROS": "Konec",
        "Promet_ROS": "Promet (ROS)",
    })
    st.dataframe(show, use_container_width=True, hide_index=True)
    st.caption(
        "To so lahko rezervacije iz drugega meseca/obdobja statementa, storno rezervacije, "
        "ali napačno/manjkajoče vnesena referenca v ROS."
    )
